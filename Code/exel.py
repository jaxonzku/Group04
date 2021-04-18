import MySQLdb
import  openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jackson\Desktop\Group04\Code\userfiles\userpass.xlsx')  
sheet = wb.active
len=sheet.max_row
print(len)


db=MySQLdb.connect("localhost","root","","college_app")
insertrec=db.cursor()

for i in range(2,len):
    x1=sheet.cell(row=i,column=1)
    x2=sheet.cell(row=i,column=2)
    x3=sheet.cell(row=i,column=3)

   # print (x1.value , x2.value)
   #ivde ini onnum chyarth...........
    string_value="'"+str(x1.value)+"'"+","+"'"+str(x2.value)+"'"+","+"'"+str(x3.value)+"'"
    print(string_value)
    sqlquery="insert into college_01(username,password,id) values ("+string_value+")"
    insertrec.execute(sqlquery)
db.commit()
print("donneeee..............")
db.close()