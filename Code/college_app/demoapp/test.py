import  openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jackson\Desktop\Group04\Code\userfiles\userpass.xlsx')  
sheet = wb.active

len=sheet.max_row
print(len)



for i in range(2,len):
    x1=sheet.cell(row=i,column=1)
    x2=sheet.cell(row=i,column=2)
    print (x1.value , x2.value)
