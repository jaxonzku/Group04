#C:\Users\jackson\AppData\Local\Programs\Python\Python39\python.exe

import os,re
from os import path
from pathlib import Path
import subprocess
import MySQLdb
import  openpyxl

####################################-------------------FUNCTION SECTIONNN----------##############################################################

constants=['inp1','inp2','inp3']
try:
    os.remove(r"C:\xampp\htdocs\college_app\demoapp\build\app\outputs\apk\release\app-release.apk")
except :
    print("file not present")


print ("1st test File exists:"+str(path.exists(r'C:\xampp\htdocs\college_app\demoapp\build\app\outputs\apk\release\app-release.apk')))





def obtainer():
    data=[]
    fname=r'C:\xampp\htdocs\details.txt'
    with open (fname, "r") as myfile:
        for line in myfile:
            #print(line)
            data.append(line.rstrip('\n'))
    print(data)
    return data



def replacer(inp,inp2):#dont get confused between inp and inp2,randum ipoll correctaanu.
    directory = os.listdir(r"C:\xampp\htdocs\college_app\demoapp\lib")
    os.chdir(r'C:\xampp\htdocs\college_app\demoapp\lib')
    for  file in directory:
        if Path(file).is_file():
                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile(inp2)
                read_file = regex.sub(inp, read_file)
                write_file = open(file, 'w')
                #print(write_file)
                write_file.write(read_file)
                #print("changed :",read_file)

def reverser(inp,inp2):#dont get confused between inp and inp2,randum ipoll correctaanu.
    directory = os.listdir(r"C:\xampp\htdocs\college_app\demoapp\lib")
    os.chdir(r'C:\xampp\htdocs\college_app\demoapp\lib')
    for  file in directory:
        if Path(file).is_file():
                #print(inp,inp2)
                
                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile(inp)
                read_file = regex.sub(inp2, read_file)
                write_file = open(file, 'w')
                #print(write_file)
                write_file.write(read_file)

def dbcreater():
   # wb = openpyxl.load_workbook(r'C:\Users\jackson\Desktop\Group04\Code\userfiles\userpass.xlsx')   #for my system
    wb = openpyxl.load_workbook(r'C:\xampp\htdocs\code\userfiles') 
    sheet = wb.active
    len=sheet.max_row
    print(len)

 # make change in here ....server db details..
    db=MySQLdb.connect("localhost","root","","college_app")
    insertrec=db.cursor()

    for i in range(2,len):
        x1=sheet.cell(row=i,column=1)
        x2=sheet.cell(row=i,column=2)
        x3=sheet.cell(row=i,column=3)

    # print (x1.value , x2.value)
    #ivde ini onnum chyarth...........
        string_value="'"+str(x1.value)+"'"+","+"'"+str(x2.value)+"'"+","+"'"+str(x3.value)+"'"
        print(string_value)
        sqlquery="insert into college_01(username,password,id) values ("+string_value+")"
        insertrec.execute(sqlquery)
    db.commit()
    print("donneeee..............")
    db.close()
    

#########################--ONLY EXECUTIONS HERE--############################################
dbcreater()



data=obtainer()
#print(data,"2")
#print(len(data),len(constants))
for i in range(len(data)):
    #print(i)
    replacer(str(data[i]),str(constants[i]))



process=subprocess.Popen(["powershell","flutter build apk --build-name=1.0.1 --build-number=1"],stdout=subprocess.PIPE);
result=process.communicate()[0]
print (result,)



data=obtainer()#reverse values on input

for i in range(len(data)):
    #print(i)
    reverser(str(data[i]),str(constants[i]))

print ("File exists:"+str(path.exists(r'C:\xampp\htdocs\college_app\demoapp\build\app\outputs\apk\release\app-release.apk')))



