#C:\Users\jackson\AppData\Local\Programs\Python\Python39\python.exe
import os,re
from os import path
from pathlib import Path
import subprocess
import  openpyxl
import json
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from firebase import firebase
import shutil
#import MySQLdb


file = open(r"C:\xampp\htdocs\log.txt","r+")
file.truncate(0)


def logger(texts):
    f= open(r"C:\xampp\htdocs\log.txt","a")
    f.write(texts+'\n')






collegename=''
global college_appname
####################################-------------------FUNCTION SECTIONNN----------##############################################################

constants=['input01','input02','input03','input04','input05','input06','input07']

logt="STEP 1 : Removing OLD apk file from APP directory"
logger(logt)
try:
    os.remove(r"C:\xampp\htdocs\app\build\app\outputs\apk\debug\app-debug.apk")
    logt="STEP 1 : success"
    logger(logt)
except :
    print("file apk not present")
    logt="STEP 1 : file apk not present "
    logger(logt)



logt="STEP 2 : Removing OLD apk file from WEB directory"
logger(logt)
try:
    os.remove(r"C:\xampp\htdocs\apk\app-debug.apk")
    logt="STEP 2 : success"
    logger(logt)
except :
    print("file apk not present")
    logt="STEP 2 : file apk not present"
    logger(logt)

  

logt="STEP 3 : Removing OLD college_image file from APP directory"
logger(logt)
try:
    os.remove(r"C:\xampp\htdocs\app\assets\images\college.jpg")
    logt="STEP 3 : success"
    logger(logt)

except:
    print("collage_image not present in app")
    logt="STEP 3 : collage_image not present in app"
    logger(logt)


logt="STEP 4:  Removing OLD college_logo file from APP directory"
logger(logt)

try:
    os.remove(r"C:\xampp\htdocs\app\assets\images\logo.jpg")
except:
    print("logo  not present in app")
    logt="STEP 4 : collage_logo not present in app"
    logger(logt)

logt="STEP 5:  checking if apk exists"
logger(logt)
print ("1st test File exists:"+str(path.exists(r'C:\xampp\htdocs\app\build\app\outputs\apk\debug\app-debug.apk')))
logger(("STEP 5 : "+str(path.exists(r'C:\xampp\htdocs\app\build\app\outputs\apk\debug\app-debug.apk'))))




def obtainer():
    data=[]
    fname=r'C:\xampp\htdocs\details.txt'
    with open (fname, "r") as myfile:
        for line in myfile:
            #print(line)
            data.append(line.rstrip('\n'))
    
    college_appname = data[1]

    print(data)
    return data



def replacer(inp,inp2,inp3):                                                                                                                #dont get confused between inp and inp2,randum ipoll correctaanu.
    directory = os.listdir(r"C:\xampp\htdocs\app\lib")
    os.chdir(r'C:\xampp\htdocs\app\lib')
    for  file in directory:
        if Path(file).is_file():
                if(file=='constants.dart'):
                    open_file = open(file,'r')
                    read_file = open_file.read()
                    regex = re.compile(inp2)
                    read_file = regex.sub(inp, read_file)
                    write_file = open(file, 'w')
                    write_file.write(read_file)

    directory = os.listdir(r"C:\xampp\htdocs\app\android\app\src\main")
    os.chdir(r'C:\xampp\htdocs\app\android\app\src\main')
    for  file in directory:
        if Path(file).is_file():

            if(file=='AndroidManifest.xml'):
                print("found xmll")

                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile("input0name0")
                read_file = regex.sub(inp3, read_file)
                write_file = open(file, 'w')
                write_file.write(read_file)
    directory = os.listdir(r'C:\xampp\htdocs\app\android\app')
    os.chdir(r'C:\xampp\htdocs\app\android\app')
    for  file in directory:
        if Path(file).is_file():
            if(file=='build.gradle'):
                print("found build.gradle")

                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile("@packname@")
                read_file = regex.sub(inp3, read_file)
                write_file = open(file, 'w')
                write_file.write(read_file)

            if(file=='google-services.json'):
                print("found build.gradle")

                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile("@packname@")
                read_file = regex.sub(inp3, read_file)
                write_file = open(file, 'w')
                write_file.write(read_file)



def reverser(inp,inp2,inp3):                                                                                                                            #dont get confused between inp and inp2,randum ipoll correctaanu.
    directory = os.listdir(r"C:\xampp\htdocs\app\lib")
    os.chdir(r'C:\xampp\htdocs\app\lib')
    content=[   "String collegename = 'input01';",
                "String appname = 'input02';",
                "String about = 'input03';",
                "String madebecause = 'input04';",
                "String library = 'input05';",
                "String collegeweb = 'input06';",
                "String chatusers = appname +'messages';",
                "String messages = appname +'messages';",
                "String userdata = appname +'users';"
            ]
    for  file in directory:
        if Path(file).is_file():
                if(file=='constants.dart'):
                    
                    write_file = open(file, 'w')
                    for i in content:
                        write_file.write(i+'\n')
    
    directory = os.listdir(r"C:\xampp\htdocs\app\android\app\src\main")
    os.chdir(r'C:\xampp\htdocs\app\android\app\src\main')
    name="input0name0"
    for  file in directory:
        if Path(file).is_file():

            if(file=='AndroidManifest.xml'):
                print("found xmll")
                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile(inp3)
                read_file = regex.sub(name, read_file)
                write_file = open(file, 'w')
                write_file.write(read_file)
    directory = os.listdir(r'C:\xampp\htdocs\app\android\app')
    os.chdir(r'C:\xampp\htdocs\app\android\app')
    name="@packname@"
    for  file in directory:
        if Path(file).is_file():

            if(file=='build.gradle'):
                print("replaced gradle")
                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile(inp3)
                read_file = regex.sub(name, read_file)
                write_file = open(file, 'w')
                write_file.write(read_file)


            if(file=='google-services.json'):
                print("replaced json")
                open_file = open(file,'r')
                read_file = open_file.read()
                regex = re.compile(inp3)
                read_file = regex.sub(name, read_file)
                write_file = open(file, 'w')
                write_file.write(read_file)
        
        
        

def dbcreater():
    wb = openpyxl.load_workbook(r'C:\xampp\htdocs\code\userfiles') 
    sheet = wb.active
    len=sheet.max_row
    print(len)

 # make change in here ....server db details..
    db=MySQLdb.connect("localhost","root","","college_app")
    insertrec=db.cursor()

    for i in range(2,len):
        x1=sheet.cell(row=i,column=1)
        x2=sheet.cell(row=i,column=2)
        x3=sheet.cell(row=i,column=3)

    # print (x1.value , x2.value)
    #ivde ini onnum chyarth...........
        string_value="'"+str(x1.value)+"'"+","+"'"+str(x2.value)+"'"+","+"'"+str(x3.value)+"'"
        print(string_value)
        sqlquery="insert into college_01(username,password,id) values ("+string_value+")"
        insertrec.execute(sqlquery)
    db.commit()
    print("donneeee..............")
    db.close()

def firebase_function(firebase,college_appname,key):
    
    df2 = pd.read_excel(r'C:\xampp\htdocs\exceluploads\data.xlsx')
    firebase = firebase.FirebaseApplication('https://collegeapp-e02e4-default-rtdb.firebaseio.com', None)
    for i in df2.index:
        a1=str(df2['regno'][i])
        a2=str(df2['branch'][i])
        a3=str(df2['email'][i])
        a4=str(df2['fullname'][i])
        a6=str(df2['rollno'][i])
        result = firebase.patch(str(college_appname)+'/'+a1,
                 {'branch': a2, 'email':a3 ,'fullName':a4,
                 'rollNo':a6,'signedin':'F'})
        result2 = firebase.patch(str(college_appname)+'/validation',
                   {'key': key})


def image_mover():
    try:
        shutil.move(r'C:\xampp\htdocs\images\college.jpg',r'C:\xampp\htdocs\app\assets\images')
        shutil.move(r'C:\xampp\htdocs\images\logo.jpg',r'C:\xampp\htdocs\app\assets\images')
        print("college_image moved")
        logger("STEP 6  : success")
    except:
        logger("STEP 6  : some error moving")
        print("some error at image mover")

def apk_mover():
    try:
        shutil.move(r'C:\xampp\htdocs\app\build\app\outputs\apk\debug\app-debug.apk',r'C:\xampp\htdocs\apk')
        print("apk moved")
    except:
        print("some error at apk mover")

    

#########################--ONLY EXECUTIONS HERE--############################################
#dbcreater() ----not used because firebase is used.


logt="STEP 6 : moving new images to app directory"
logger(logt)

image_mover()



logt="STEP 7 : obtaining data from details.txt and replacing"
logger(logt)

data=obtainer() #used to acces data from text file
for i in range(len(data)):
    replacer(str(data[i]),str(constants[i]),str(data[1]))  

college_appname=data[1]

logt="STEP 8 : calling firebase function"
logger(logt)
firebase_function(firebase,college_appname,str(data[6]))


#from here powrshell exec......mipmap generator also
logt="STEP 9 : logo generating in flutter"
logger(logt)

process1=subprocess.Popen(["powershell","flutter pub run flutter_launcher_icons:main"],stdout=subprocess.PIPE,);
result1=process1.communicate()[0]
print (result1)

logt="STEP 10 : apk generating in flutter"
logger(logt)

process2=subprocess.Popen(["powershell","flutter build apk --debug"],stdout=subprocess.PIPE);
result2=process2.communicate()[0]
print (result2)


logt="STEP 11 : reversing variables"
logger(logt)
data=obtainer()#reverse values on input

for i in range(len(data)):
    reverser(str(data[i]),str(constants[i]),str(data[1]))
print ("File exists:"+str(path.exists(r'C:\xampp\htdocs\app\build\app\outputs\apk\debug\app-debug.apk')))


logt="STEP 12 : moving apk"
logger(logt)
apk_mover()

